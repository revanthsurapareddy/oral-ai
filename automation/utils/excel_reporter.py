import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from automation.config.config import Config
from automation.utils.logger import AutomationLogger

class ExcelReporter:
    @staticmethod
    def generate_all_excel_reports(results, summary_stats):
        Config.init_dirs()
        logger = AutomationLogger.get_logger()
        logger.info("Generating Enterprise Excel Reports...")

        main_report_path = os.path.join(Config.EXCEL_REPORTS_DIR, "Automation_Test_Report.xlsx")
        passed_report_path = os.path.join(Config.EXCEL_REPORTS_DIR, "Passed_Test_Cases.xlsx")
        failed_report_path = os.path.join(Config.EXCEL_REPORTS_DIR, "Failed_Test_Cases.xlsx")
        summary_report_path = os.path.join(Config.EXCEL_REPORTS_DIR, "Summary_Report.xlsx")

        # 1. Main Comprehensive Report (6 Sheets)
        wb = openpyxl.Workbook()
        
        # Styles
        header_fill = PatternFill(start_color="1F2C3B", end_color="1F2C3B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        pass_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        fail_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
        skip_fill = PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid")

        # Sheet 1: Executed Test Cases
        ws1 = wb.active
        ws1.title = "Executed Test Cases"
        headers1 = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority", "Details / Error"]
        ws1.append(headers1)
        
        for r in results:
            ws1.append([
                r["test_id"], r["module"], r["test_name"], r["status"],
                round(r.get("execution_time", 0.05), 3), r.get("priority", "P2"),
                r.get("failure_reason", "")
            ])

        # Sheet 2: Passed Tests
        ws2 = wb.create_sheet(title="Passed Tests")
        ws2.append(["Test ID", "Module", "Test Name", "Execution Time (s)", "Priority"])
        for r in results:
            if r["status"] == "PASS":
                ws2.append([r["test_id"], r["module"], r["test_name"], round(r.get("execution_time", 0.05), 3), r.get("priority", "P2")])

        # Sheet 3: Failed Tests
        ws3 = wb.create_sheet(title="Failed Tests")
        ws3.append(["Test ID", "Module", "Test Name", "Failure Reason", "Stack Trace / Log", "Priority"])
        for r in results:
            if r["status"] == "FAIL":
                ws3.append([r["test_id"], r["module"], r["test_name"], r.get("failure_reason", "Assertion Failed"), r.get("stack_trace", "N/A"), r.get("priority", "P1")])

        # Sheet 4: Skipped Tests
        ws4 = wb.create_sheet(title="Skipped Tests")
        ws4.append(["Test ID", "Module", "Test Name", "Skip Reason", "Priority"])
        for r in results:
            if r["status"] == "SKIP":
                ws4.append([r["test_id"], r["module"], r["test_name"], r.get("failure_reason", "Precondition not met"), r.get("priority", "P3")])

        # Sheet 5: Execution Metrics
        ws5 = wb.create_sheet(title="Execution Metrics")
        ws5.append(["Metric", "Value"])
        ws5.append(["Total Executed", summary_stats["total"]])
        ws5.append(["Passed", summary_stats["passed"]])
        ws5.append(["Failed", summary_stats["failed"]])
        ws5.append(["Skipped", summary_stats["skipped"]])
        ws5.append(["Pass Percentage", f"{summary_stats['pass_rate']}%"])
        ws5.append(["Execution Duration (s)", summary_stats["duration"]])
        ws5.append(["Target Environment", summary_stats["base_url"]])

        # Sheet 6: Defect Summary
        ws6 = wb.create_sheet(title="Defect Summary")
        ws6.append(["Defect ID", "Module", "Test ID", "Severity", "Description"])
        defect_id = 101
        for r in results:
            if r["status"] == "FAIL":
                ws6.append([f"DEF-{defect_id}", r["module"], r["test_id"], r.get("priority", "High"), r.get("failure_reason", "Unexpected failure")])
                defect_id += 1

        # Format sheets
        for sheet in wb.worksheets:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

        wb.save(main_report_path)
        logger.info("Saved Automation_Test_Report.xlsx")

        # 2. Passed Test Cases Workbook
        wb_pass = openpyxl.Workbook()
        wsp = wb_pass.active
        wsp.title = "Passed Tests"
        wsp.append(["Test ID", "Module", "Test Name", "Status", "Priority"])
        for r in results:
            if r["status"] == "PASS":
                wsp.append([r["test_id"], r["module"], r["test_name"], "PASS", r.get("priority", "P2")])
        wb_pass.save(passed_report_path)

        # 3. Failed Test Cases Workbook
        wb_fail = openpyxl.Workbook()
        wsf = wb_fail.active
        wsf.title = "Failed Tests"
        wsf.append(["Test ID", "Module", "Test Name", "Failure Reason", "Priority"])
        for r in results:
            if r["status"] == "FAIL":
                wsf.append([r["test_id"], r["module"], r["test_name"], r.get("failure_reason", ""), r.get("priority", "P1")])
        wb_fail.save(failed_report_path)

        # 4. Summary Report Workbook
        wb_sum = openpyxl.Workbook()
        wss = wb_sum.active
        wss.title = "Summary Report"
        wss.append(["Metric", "Count / Value"])
        wss.append(["Total Executed", summary_stats["total"]])
        wss.append(["Passed", summary_stats["passed"]])
        wss.append(["Failed", summary_stats["failed"]])
        wss.append(["Skipped", summary_stats["skipped"]])
        wss.append(["Pass Percentage", f"{summary_stats['pass_rate']}%"])
        wss.append(["Execution Duration (s)", summary_stats["duration"]])
        wb_sum.save(summary_report_path)

        logger.info("All 4 Excel reports generated successfully.")
